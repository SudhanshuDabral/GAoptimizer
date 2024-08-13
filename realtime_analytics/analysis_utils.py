
import numpy as np
import logging
import pandas as pd


logger = logging.getLogger(__name__)

def log_message(level, message):
    logger.log(level,f"[realtime-analytics] {message}")

def calculate_slope(x, y):
    try:
        return np.polyfit(x, y, 1)[0]
    except Exception as e:
        log_message(logging.ERROR, f"Error calculating slope: {str(e)}")
        raise

def analyze_window(window):
    try:
        time = window['Time Seconds']
        tp = window['Treating Pressure']
        sr = window['Slurry Rate']
        ppc = window['PPC']
        bh_prop_mass = window['BH Prop Mass']
        
        tp_slope = calculate_slope(time, tp)
        sr_slope = calculate_slope(time, sr)
        ppc_change = ppc.iloc[-1] - ppc.iloc[0]
        
        condition1 = tp_slope < 0 and sr_slope > 0
        condition2 = tp_slope < 0 and abs(sr_slope) < 1e-6
        
        # New analysis
        min_tp_idx = tp.idxmin()
        max_tp_idx = tp.idxmax()
        min_tp = tp.loc[min_tp_idx]
        max_tp = tp.loc[max_tp_idx]
        sr_at_min_tp = sr.loc[min_tp_idx]
        sr_at_max_tp = sr.loc[max_tp_idx]
        bh_at_min_tp = bh_prop_mass.loc[min_tp_idx]
        bh_at_max_tp = bh_prop_mass.loc[max_tp_idx]
        
        new_analysis = {
            'Min Treating Pressure': min_tp,
            'Max Treating Pressure': max_tp,
            'Slurry Rate @ Min TP': sr_at_min_tp,
            'Slurry Rate @ Max TP': sr_at_max_tp,
            'BH Prop Mass @ Min TP': bh_at_min_tp,
            'BH Prop Mass @ Max TP': bh_at_max_tp,
            'Delta Treating Pressure': max_tp - min_tp,
            'Delta Slurry Rate': sr_at_max_tp - sr_at_min_tp,
            'Delta BH Prop Mass': bh_at_max_tp - bh_at_min_tp
        }
        
        return (condition1 or condition2), ppc_change, new_analysis
    except Exception as e:
        log_message(logging.ERROR, f"Error analyzing window: {str(e)}")
        raise

def detect_leakoff_periods(data):
    try:
        leakoff_periods = []
        in_leakoff = False
        start_time = None
        found_first_slurry = False

        for index, row in data.iterrows():
            if row['Slurry Rate'] > 0:
                found_first_slurry = True
            if found_first_slurry:
                if row['Slurry Rate'] == 0 and row['Treating Pressure'] > 0:
                    if not in_leakoff:
                        # Start of a leakoff period
                        in_leakoff = True
                        start_time = row['datetime']
                else:
                    if in_leakoff:
                        # End of a leakoff period
                        in_leakoff = False
                        end_time = row['datetime']
                        leakoff_periods.append((start_time, end_time))
                        start_time = None

        if in_leakoff:
            # If the data ends while still in a leakoff period
            leakoff_periods.append((start_time, data['datetime'].iloc[-1]))

        return leakoff_periods
    except Exception as e:
        log_message(logging.ERROR, f"Error detecting leakoff periods: {str(e)}")
        raise

def analyze_data(df, window_size):
    try:
        total_time = df['Time Seconds'].max() - df['Time Seconds'].min()
        window_step = (window_size + 1) // 2
        total_windows = max(1, int((total_time - window_size) / window_step) + 1)
        
        event_count = 0
        total_ppc_change = 0
        event_windows = []
        new_analysis_results = []
        
        leakoff_periods = detect_leakoff_periods(df)
        
        log_message(logging.INFO, f"Starting analysis with {total_windows} windows")
        
        for i in range(total_windows):
            start_time = df['Time Seconds'].min() + i * window_step
            end_time = start_time + window_size
            window = df[(df['Time Seconds'] >= start_time) & (df['Time Seconds'] < end_time)]
            
            if len(window) > 0:
                event, ppc_change, new_analysis = analyze_window(window)
                new_analysis['Window Start'] = start_time
                new_analysis['Window End'] = end_time
                new_analysis_results.append(new_analysis)
                
                if event:
                    is_in_leakoff = any(start <= window['datetime'].iloc[0] <= end for start, end in leakoff_periods)
                    if not is_in_leakoff:
                        event_count += 1
                        total_ppc_change += ppc_change
                        event_windows.append((start_time, end_time))
            
            if i % 100 == 0:
                log_message(logging.INFO, f"Processed {i}/{total_windows} windows")
        
        log_message(logging.INFO, f"Analysis completed: {event_count} events found out of {total_windows} windows")
        return total_windows, event_count, total_ppc_change, event_windows, leakoff_periods, new_analysis_results
    except Exception as e:
        log_message(logging.ERROR, f"Error analyzing data: {str(e)}")
        raise

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def export_to_excel(combined_data, well_name):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Write summary information
    ws['A1'] = f"Rolling IR Data for {well_name}"
    ws['A2'] = f"Total data points: {len(combined_data)}"
    ws['A3'] = f"Unique stages: {combined_data['stage'].nunique()}"

    # Create sheets for each stage
    for stage in combined_data['stage'].unique():
        stage_data = combined_data[combined_data['stage'] == stage]
        ws = wb.create_sheet(title=f"Stage {stage}")
        for r in dataframe_to_rows(stage_data[['normalized_time', 'rolling_IR']], index=False, header=True):
            ws.append(r)

    # Save the workbook to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def calculate_slope(x, y):
    n = len(x)
    if n < 2:
        return np.nan
    sum_x = np.sum(x)
    sum_y = np.sum(y)
    sum_xy = np.sum(x * y)
    sum_xx = np.sum(x * x)
    denominator = n * sum_xx - sum_x * sum_x
    if denominator == 0:
        return np.nan
    slope = (n * sum_xy - sum_x * sum_y) / denominator
    return slope

def calculate_rolling_ir(df, stage):
    log_message(logging.INFO, f"Starting rolling IR calculation for stage {stage}")
    log_message(logging.INFO, f"DataFrame shape: {df.shape}")
    log_message(logging.INFO, f"Columns: {df.columns}")
    
    # Check if required columns exist
    required_columns = ['time_seconds', 'slurry_rate', 'treating_pressure']
    for col in required_columns:
        if col not in df.columns:
            log_message(logging.ERROR, f"Required column '{col}' not found in DataFrame for stage {stage}")
            return pd.DataFrame(columns=['normalized_time', 'rolling_IR'])

    # Aggregate data by second
    df_agg = df.groupby('time_seconds').agg({
        'slurry_rate': 'mean',
        'treating_pressure': 'mean'
    }).reset_index()

    # Find the first index where slurry_rate > 0 and stays > 0 for 15 seconds
    start_index = None
    for i in range(len(df_agg) - 15):  # -15 to ensure we have 15 seconds to check
        if df_agg.loc[i, 'slurry_rate'] > 0 and all(df_agg.loc[i:i+15, 'slurry_rate'] > 0):
            start_index = i
            break

    if start_index is None:
        log_message(logging.WARNING, f"No valid start point found for stage {stage}")
        return pd.DataFrame(columns=['normalized_time', 'rolling_IR'])

    start_time = df_agg.loc[start_index, 'time_seconds']
    
    # Process data for the next 15 minutes (900 seconds) from this starting point
    end_time = start_time + 900
    df_filtered = df_agg[(df_agg['time_seconds'] >= start_time) & (df_agg['time_seconds'] <= end_time)].copy()
    
    # Find the index where slurry_rate first exceeds 5 within this window
    ir_start_indices = df_filtered[df_filtered['slurry_rate'] > 5].index
    if len(ir_start_indices) == 0:
        log_message(logging.WARNING, f"No data points where slurry_rate > 5 in 15-minute window for stage {stage}")
        return pd.DataFrame(columns=['normalized_time', 'rolling_IR'])
    
    ir_start_index = ir_start_indices[0]
    
    # Calculate cumulative sums from the ir_start_index
    df_filtered.loc[ir_start_index:, 'cum_slurry_rate'] = df_filtered.loc[ir_start_index:, 'slurry_rate'].cumsum()
    df_filtered.loc[ir_start_index:, 'cum_treating_pressure'] = df_filtered.loc[ir_start_index:, 'treating_pressure'].cumsum()
    
    # Calculate rolling_IR (slope)
    num_rows = len(df_filtered.loc[ir_start_index:])
    rolling_ir = np.full(num_rows, np.nan)
    
    for i in range(num_rows):
        x = df_filtered.loc[ir_start_index:ir_start_index+i+1, 'cum_treating_pressure'].values
        y = df_filtered.loc[ir_start_index:ir_start_index+i+1, 'cum_slurry_rate'].values
        rolling_ir[i] = calculate_slope(x, y)
    
    df_filtered.loc[ir_start_index:, 'rolling_IR'] = rolling_ir
    
    # Normalize time_seconds to start from 0 for each stage
    df_filtered.loc[ir_start_index:, 'normalized_time'] = df_filtered.loc[ir_start_index:, 'time_seconds'] - df_filtered.loc[ir_start_index, 'time_seconds']
    
    result = df_filtered.loc[ir_start_index:, ['normalized_time', 'rolling_IR']].copy()
    result['normalized_time'] = result['normalized_time'].astype(int)
    
    log_message(logging.INFO, f"Filtered DataFrame shape: {result.shape}")
    
    return result
